#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsm2spec extract.py
Excel(.xlsm)を解析し、AIが仕様化分析に使える中間成果物(Markdown)を抽出する。

出力構成:
  <out>/
    vba/<module>.txt           ... VBAモジュールごとの完全ソース
    00_workbook_overview.md    ... ファイル情報・ワークブック構成・定義名
    10_sheet_list.md           ... シート一覧・役割推定
    20_vba_summary.md          ... モジュール/プロシージャ/呼び出し/グローバル/DB/メッセージ
    30_buttons.md              ... ボタン→マクロ割当一覧
    40_cross_references.md     ... VBAとシート/DBのクロス参照
    sheets/<n>_<name>.md       ... シートごとの構造詳細
    50_db_schema.md            ... Access DBスキーマ検証 (--db 指定時)

依存: openpyxl, oletools (pip install openpyxl oletools)
DB解析時: access_parser (pip install access_parser)
"""
import argparse
import io
import os
import re
import sys
import warnings
from collections import Counter, defaultdict
from datetime import datetime
from xml.etree import ElementTree as ET

MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
RELS_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
MC_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

VBA_KEYWORDS = {
    "ByVal", "ByRef", "Dim", "As", "Option", "Explicit", "Public", "Private",
    "Static", "Const", "Set", "Call", "Sub", "Function", "End", "If", "Then",
    "Else", "ElseIf", "For", "Next", "To", "Step", "While", "Wend", "Do",
    "Loop", "Until", "Select", "Case", "Exit", "With", "ReDim", "Preserve",
    "Type", "Enum", "Declare", "Lib", "Let", "Get", "Property", "On", "GoTo",
    "GoSub", "Return", "Resume", "Error", "Each", "In", "Is", "Not", "And",
    "Or", "Xor", "Mod", "New", "Nothing", "True", "False", "Null", "Empty",
    "Mid", "Left", "Right", "Len", "Trim", "LTrim", "RTrim", "UCase", "LCase",
    "Str", "CStr", "Val", "CInt", "CLng", "CDbl", "CSng", "CDec", "CBool",
    "CByte", "CChar", "CDate", "CType", "Abs", "Int", "Fix", "Round", "Sqr",
    "Exp", "Log", "Sin", "Cos", "Tan", "Atn", "Rnd", "Randomize", "Date",
    "Time", "Now", "Year", "Month", "Day", "Hour", "Minute", "Second", "Weekday",
    "DateSerial", "TimeSerial", "DateValue", "TimeValue", "DateAdd", "DateDiff",
    "DatePart", "Format", "FormatDateTime", "MsgBox", "InputBox", "IsNull",
    "IsNumeric", "IsDate", "IsEmpty", "IsArray", "IsObject", "IsError",
    "Array", "UBound", "LBound", "LBound", "Filter", "Join", "Split", "Replace",
    "InStr", "InStrRev", "StrConv", "Asc", "AscW", "Chr", "ChrW", "Space",
    "String", "TypeName", "VarType", "Debug", "Print", "Stop", "ScreenUpdating",
    "DisplayAlerts", "Application", "ActiveWorkbook", "ActiveSheet", "Workbooks",
    "Worksheets", "Sheets", "Range", "Cells", "Rows", "Columns", "Select",
    "Selection", "ActiveCell", "UsedRange", "Rows", "Name", "Names", "Font",
    "Interior", "Borders", "WrapText", "NumberFormat", "ColumnWidth", "RowHeight",
    "Merge", "UnMerge", "AutoFilter", "AutoFilterMode", "Sort", "Order1",
    "Order2", "Order3", "xlAscending", "xlDescending", "Find", "FindFirst",
    "FindNext", "FindLast", "Edit", "AddNew", "Update", "Delete", "Refresh",
    "OpenRecordset", "OpenDatabase", "Recordset", "Database", "Connection",
    "dbOpenDynaset", "dbOpenSnapshot", "Execute", "EOF", "BOF", "MoveFirst",
    "MoveLast", "MoveNext", "MovePrevious", "RecordCount", "NoMatch", "Fields",
    "ListIndex", "RowSource", "ColumnHeads", "ColumnCount", "ColumnWidths",
    "ListWidth", "AddItem", "RemoveItem", "Clear", "SetFocus", "Show", "Unload",
    "Load", "Hide", "Visible", "Value", "Text", "Caption", "Enabled", "Locked",
    "TabIndex", "MultiLine", "PasswordChar", "KeyDown", "KeyPress", "KeyUp",
    "Change", "Click", "DblClick", "BeforeDoubleClick", "Cancel", "Target",
    "vbUpperCase", "vbNarrow", "vbWide", "vbYes", "vbNo", "vbOK", "vbCancel",
    "vbInformation", "vbExclamation", "vbQuestion", "vbCritical", "MsgBoxStyle",
    "MsgBoxResult", "vbOKCancel", "vbYesNo", "vbAbortRetryIgnore", "vbYesNoCancel",
    "vbQuestion", "Chr", "OpenDatabase", "ThisWorkbook", "Workbook", "Worksheet",
    "CommandButton", "TextBox", "Label", "ComboBox", "ListBox", "Frame", "UserForm",
    "CheckBox", "OptionButton", "Image", "TabStrip", "MultiPage", "ScrollBar",
    "SpinButton", "ToggleButton", "Cancel", "MousePointer", "List", "BoundColumn",
    "TextColumn", "ControlSource", "TakeFocusOnClick", "LockText",
}

EVENT_SUFFIXES = (
    "Click", "Change", "DblClick", "BeforeDoubleClick", "BeforeRightClick",
    "Activate", "Deactivate", "Initialize", "Terminate", "AfterUpdate",
    "BeforeUpdate", "Enter", "Exit", "GotFocus", "LostFocus", "KeyDown",
    "KeyPress", "KeyUp", "MouseDown", "MouseMove", "MouseUp", "Scroll",
    "SpinDown", "SpinUp", "DragDrop", "DragOver", "DropButtonClick", "Close",
    "Open", "QueryClose", "Layout", "Resize", "StartDrag", "BeginDrag",
    "EndDrag", "Refresh", "SelectionChange", "Calculate", "Change", "RowChange",
)


def _bootstrap_venv():
    """pip無し環境向け: スキル同梱の .venv が無く依存も未導入なら、自身をvenvのPythonで再実行する。
    install.sh が pip を導入できなかった場合に `$DEST/.venv` を作るため、
    `python3 <スキル>/scripts/extract.py` の呼び出し方のままで動作させるための仕組み。"""
    try:
        import openpyxl  # noqa: F401
        import oletools  # noqa: F401
        return
    except Exception:
        pass
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for cand in (os.path.join(script_dir, ".venv"),
                 os.path.join(script_dir, "..", ".venv")):
        py = os.path.join(cand, "bin", "python3")
        if os.path.isfile(py):
            os.execv(py, [py] + sys.argv)


_bootstrap_venv()


def log(msg):
    sys.stderr.write("[extract] %s\n" % msg)


def read_zip_xml(zpath, entry):
    """zip内のXMLエントリをElementTreeとして読む(無ければNone)。"""
    import zipfile
    try:
        with zipfile.ZipFile(zpath) as z:
            with z.open(entry) as f:
                return ET.fromstring(f.read())
    except (KeyError, ET.ParseError):
        return None


def sheet_order(zpath):
    """workbook.xmlから (sheet名, sheetId, rId) の順序リストを返す。"""
    root = read_zip_xml(zpath, "xl/workbook.xml")
    if root is None:
        return []
    out = []
    for sh in root.findall(f"{MAIN_NS}sheets/{MAIN_NS}sheet"):
        out.append((sh.get("name"), sh.get("sheetId"), sh.get("{%s}id" % RELS_NS)))
    return out


def sheet_zip_file(zpath, rels_entry, r_id):
    """rIdから xl/worksheets/sheetN.xml を特定する。"""
    rels = read_zip_xml(zpath, rels_entry)
    if rels is None:
        return None
    for rel in rels.findall(f"{RELS_NS}Relationship"):
        if rel.get("Id") == r_id:
            tgt = rel.get("Target", "")
            if not tgt.startswith("/"):
                base = rels_entry.rsplit("/", 1)[0] + "/"
                tgt = base + tgt
            return tgt.replace("/xl/", "xl/")
    return None


def buttons_map(zpath, sheet_entries):
    """シートXMLに埋め込まれた macro="[0]!名前" を収集。"""
    import zipfile
    result = {}
    with zipfile.ZipFile(zpath) as z:
        for entry in sheet_entries:
            base = os.path.basename(entry)
            idx = re.search(r"sheet(\d+)\.xml", base)
            if not idx:
                continue
            sn = "sheet%d" % int(idx.group(1))
            try:
                content = z.read(entry).decode("utf-8", errors="replace")
            except KeyError:
                continue
            macros = re.findall(r'macro="([^"]*)"', content)
            macros = [m for m in macros if m.strip()]
            macros = [re.sub(r"^\[[0-9]+\]!", "", m) for m in macros]
            result[sn] = macros
    return result


def collect_images(zpath):
    """drawingから image/control の参照数を数える。"""
    import zipfile
    counts = Counter()
    with zipfile.ZipFile(zpath) as z:
        names = z.namelist()
    for entry in names:
        m = re.match(r"xl/drawings/drawing(\d+)\.xml$", entry)
        if not m:
            continue
        root = read_zip_xml(zpath, entry)
        if root is None:
            continue
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        for e in root.iter():
            tag = e.tag.split("}")[-1]
            if tag in ("pic", "cxnSp", "sp", "control", "graphicFrame"):
                counts["drawing%d" % int(m.group(1))] += 1
    return counts


def extract_vba(zpath):
    """olevbaでVBAを抽出。[(モジュール名, タイプ, コード)]"""
    try:
        from oletools.olevba import VBA_Parser
    except ImportError:
        log("WARN: oletoolsが未導入です。VBA解析をスキップします (pip install oletools)")
        return []
    try:
        parser = VBA_Parser(zpath)
        mods = []
        for filename, stream, vba_name, code in parser.extract_macros():
            ftype = os.path.splitext(vba_name)[1].lstrip(".") if vba_name else ""
            mods.append((vba_name, ftype, code))
        parser.close()
        return mods
    except Exception as e:
        log("WARN: VBA抽出に失敗: %s" % e)
        return []


def vba_project_streams(zpath):
    """zip内の vbaProject.bin をOLEとして開き、全ストリームのバイト列を返す。"""
    import zipfile
    try:
        import olefile
    except ImportError:
        return {}
    try:
        with zipfile.ZipFile(zpath) as z:
            data = z.read("xl/vbaProject.bin")
    except (KeyError, zipfile.BadZipFile):
        return {}
    try:
        ole = olefile.OleFileIO(io.BytesIO(data))
    except Exception:
        return {}
    streams = {}
    for entry in ole.listdir():
        try:
            streams["/".join(entry)] = ole.openstream(entry).read()
        except Exception:
            continue
    ole.close()
    return streams


def form_designer_strings(fdata):
    """MS Formsデザイナー(/f)からASCII・UTF-16LEの可読文字列を抽出する。"""
    ascii_runs = [s.decode("ascii") for s in re.findall(rb"[\x20-\x7e]{2,}", fdata)]
    utf16_runs = []
    cur = []
    for i in range(0, len(fdata) - 1, 2):
        unit = fdata[i] | (fdata[i + 1] << 8)
        if 0xD800 <= unit <= 0xDFFF:
            unit = 0
        if 0x20 <= unit <= 0x7E or 0x4E00 <= unit <= 0x9FFF or 0x3040 <= unit <= 0x30FF:
            cur.append(chr(unit))
        else:
            if len(cur) >= 1:
                utf16_runs.append("".join(cur))
            cur = []
    if cur:
        utf16_runs.append("".join(cur))
    out = []
    seen = set()
    for s in ascii_runs + utf16_runs:
        s = s.strip()
        if not s or s in seen:
            continue
        seen.add(s)
        # ASCII: Label1 のような自動生成名のみ
        if re.fullmatch(r"[A-Za-z][A-Za-z0-9]{1,}", s):
            out.append(s)
            continue
        # 日本語: 仮名を含む or 常用漢字のみ(2〜12文字) で、ASCII文字コードの誤読パターンを除外
        has_kana = any(0x3040 <= ord(ch) <= 0x30FF for ch in s)
        all_cjk = all(0x4E00 <= ord(ch) <= 0x9FFF or 0x3040 <= ord(ch) <= 0x30FF for ch in s)
        if (has_kana or all_cjk) and 1 <= len(s) <= 12 and not re.fullmatch(r"[0-9a-fA-F]{4,}", s):
            out.append(s)
    return out


def analyze_form(form_name, code, designer_strings, all_proc_names, sheet_names, code_names):
    """UserFormモジュールからコントロール一覧と設定を解析する。"""
    procs = parse_procs(code)
    # イベント → コントロール
    events = defaultdict(list)
    for p in procs:
        name = p["signature"][2]
        for suf in EVENT_SUFFIXES:
            if name.endswith("_" + suf) and not name.startswith("UserForm_"):
                ctrl = name[: -(len(suf) + 1)]
                events[ctrl].append(suf)
                break

    # コード中で識別子として現れるデザイナー文字列 → コントロール候補
    designer_ctrls = [s for s in designer_strings
                      if re.search(r"\b" + re.escape(s) + r"\b", code)
                      and not re.fullmatch(r"[A-Za-z]+[0-9]+", s)
                      and not re.fullmatch(r"[A-Za-z]{1,3}", s)]

    all_ctrls = set(events.keys()) | set(designer_ctrls)
    # コード中で 識別子.プロパティ の形で利用されるものも候補に加える
    for m in re.finditer(r"\b([A-Za-z0-9_\u3000-\u9fff]+)\.[A-Za-z]+\b", code):
        tok = m.group(1)
        if tok not in VBA_KEYWORDS and tok not in ("Range", "Cells", "ActiveSheet",
                                                    "Selection", "Sheets", "Worksheets",
                                                    "Application", "ActiveWorkbook",
                                                    "Workbooks", "ThisWorkbook", "MsgBox",
                                                    "Format", "Trim", "Str", "CStr", "CDbl",
                                                    "Now", "Year", "Month", "Null", "True",
                                                    "False", "r", "cn", "sql"):
            if re.match(r"^[A-Za-z0-9_\u3000-\u9fff]+$", tok):
                all_ctrls.add(tok)

    # コントロールごとのコード上でのプロパティ利用（Withブロック対応）
    ctrl_usage = defaultdict(set)
    for p in procs:
        body = p["body"]
        with_stack = []
        for ln in body:
            m = re.match(r"^\s*With\s+([A-Za-z0-9_\u3000-\u9fff]+)", ln)
            if m:
                with_stack.append(m.group(1))
            if ln.strip() == "End With" and with_stack:
                with_stack.pop()
            target = with_stack[-1] if with_stack else None
            for cm in re.finditer(r"\.([A-Za-z]+)", ln):
                prop = cm.group(1)
                if prop in ("End", "Then", "Else", "Next", "Do", "Loop",
                            "While", "Wend", "Select", "Case", "If", "For"):
                    continue
                if target is not None and target in all_ctrls:
                    ctrl_usage[target].add(prop)
                m2 = re.search(r"\b([A-Za-z0-9_\u3000-\u9fff]+)\." + re.escape(prop) + r"\b", ln)
                if m2 and m2.group(1) in all_ctrls:
                    ctrl_usage[m2.group(1)].add(prop)

    # 型推測
    controls = {}
    for ctrl in sorted(all_ctrls):
        evs = sorted(events.get(ctrl, []))
        usage = sorted(ctrl_usage.get(ctrl, []))
        ctype = infer_control_type(evs, usage, ctrl in designer_strings, code)
        controls[ctrl] = {"type": ctype, "events": evs, "usage": usage}

    # フォーム全体のプロシージャ（非イベント）
    form_procs = [p for p in procs if not p["signature"][2].startswith("UserForm_") and
                  not any(p["signature"][2].endswith("_" + s) for s in EVENT_SUFFIXES)]
    return {"controls": controls, "form_procs": [p["signature"][2] for p in form_procs],
            "designer": designer_strings}


def infer_control_type(events, usage, in_designer=False, code=""):
    us = set(usage)
    if us.intersection({"RowSource", "ListWidth", "ColumnCount", "ColumnHeads", "ColumnWidths", "AddItem"}):
        return "ComboBox(コンボボックス)"
    if "Click" in events and not us.intersection({"AddItem", "RowSource", "List", "ColumnCount", "ColumnHeads", "ListWidth", "Text", "Value"}):
        return "CommandButton(ボタン)"
    if "Change" in events:
        return "ComboBox(コンボボックス)"
    if "KeyDown" in events or "KeyPress" in events or "KeyUp" in events:
        return "TextBox(テキストボックス)"
    if us.intersection({"Text", "Value", "SetFocus"}) or in_designer:
        return "TextBox(テキストボックス)"
    if "Click" in events:
        return "CommandButton(ボタン)"
    return "Label/その他"


def proc_signature(line):
    """プロシージャ宣言行の解析。 (kind, name, args) or None"""
    m = re.match(
        r"^\s*(?:(Public|Private|Friend|Static)\s+)?(Sub|Function|Property\s+(?:Get|Let|Set))\s+"
        r"([A-Za-z0-9_\u3000-\u9fff]+)\s*(?:\(([^)]*)\))?\s*(?:As\s+[\w\.]+)?\s*$",
        line,
    )
    if not m:
        return None
    vis, kind, name, args = m.groups()
    return (vis or "Public", kind.strip(), name, (args or "").strip())


def parse_procs(code):
    """モジュール内のプロシージャ一覧。"""
    lines = code.splitlines()
    procs = []
    i = 0
    n = len(lines)
    while i < n:
        sig = proc_signature(lines[i])
        if sig is None:
            i += 1
            continue
        start = i
        i += 1
        depth = 0
        while i < n:
            s = lines[i].strip()
            if s.startswith(("End Sub", "End Function", "End Property")):
                break
            if s.startswith(("Sub ", "Function ", "Property ")) and not s.startswith("End "):
                depth += 1
            if s.startswith("End "):
                depth -= 1
            i += 1
        body = lines[start + 1:i]
        procs.append({"signature": sig, "start": start + 1, "body": body,
                      "end": i + 1, "line_count": (i - start + 1)})
        i += 1
    return procs


def find_proc_name(code):
    m = re.search(r"^\s*(?:Public|Private|Friend|Static\s+)?(Sub|Function|Property\s+\w+)\s+([A-Za-z0-9_\u3000-\u9fff]+)",
                  code, re.MULTILINE)
    return m.group(2) if m else ""


def first_doc_comment(body_lines, sig_line):
    """プロシージャ直前/直後のコメントを説明文として取る。"""
    docs = []
    for ln in body_lines[:6]:
        s = ln.strip()
        if s.startswith("'"):
            docs.append(s.lstrip("'").strip())
        elif s == "" or s.startswith("Call") or s.startswith("If") or s.startswith("Dim"):
            if not s.startswith("'"):
                if docs and s == "":
                    break
        if len(docs) >= 4:
            break
    return " | ".join(docs)[:200]


def analyze_proc(proc, all_proc_names, sheet_names, code_names):
    """プロシージャ本体から呼出/参照/変数/メッセージ等を解析。"""
    body = "\n".join(proc["body"])
    info = {
        "calls": set(),
        "run_calls": set(),
        "sheets": set(),
        "ranges": [],
        "db_tables": [],
        "messages": [],
        "dim_vars": [],
        "controls": set(),
    }
    # Call X と Application.Run "X" と 直接呼び出し(既知プロシージャ名)
    for m in re.finditer(r"\bCall\s+([A-Za-z0-9_\u3000-\u9fff]+)", body):
        info["calls"].add(m.group(1))
    for m in re.finditer(r'(?:Application\.)?Run\s+"([^"]+)"', body):
        info["run_calls"].add(m.group(1))
    for m in re.finditer(r"\.Run\s+\"([^\"]+)\"", body):
        info["run_calls"].add(m.group(1))
    for m in re.finditer(r"^\s*([A-Za-z0-9_\u3000-\u9fff]+)\s+", body, re.MULTILINE):
        tok = m.group(1)
        if tok in all_proc_names:
            info["calls"].add(tok)
    for name in all_proc_names:
        for m in re.finditer(r"\b" + re.escape(name) + r"\s*\(", body):
            info["calls"].add(name)
    # シート参照
    for m in re.finditer(r'(?:Sheets|Worksheets)\(["\']?([^"\')]+)["\']?\)', body):
        info["sheets"].add(m.group(1))
    for m in re.finditer(r'"([^"]+)"', body):
        if m.group(1) in sheet_names:
            info["sheets"].add(m.group(1))
    for cn in code_names:
        if cn and cn != "ThisWorkbook":
            for m in re.finditer(r"\b" + re.escape(cn) + r"\.(?:Range|Cells|Select|Activate|Visible)", body):
                info["sheets"].add(cn)
    if re.search(r"\bActiveSheet\b", body):
        info["sheets"].add("ActiveSheet")
    # 範囲参照
    for m in re.finditer(r'\.?Range\("([^"]+)"\)', body):
        info["ranges"].append(m.group(1))
    for m in re.finditer(r'\bCells\(', body):
        info["ranges"].append("Cells(...)")
    for m in re.finditer(r'\.FindFirst "([^"]+)"', body):
        info["ranges"].append("FindFirst(" + m.group(1)[:60] + ")")
    # DBテーブル
    for m in re.finditer(r'["\']([TQ]_[A-Za-z0-9_\u3000-\u9fff]+)', body):
        info["db_tables"].append(m.group(1))
    for m in re.finditer(r'\bfrom\s+([A-Za-z0-9_\u3000-\u9fff]+)\b', body, re.IGNORECASE):
        info["db_tables"].append(m.group(1))
    # MsgBoxメッセージ
    for m in re.finditer(r'MsgBox\s+"([^"]+)"', body):
        info["messages"].append(m.group(1))
    # Dim変数
    for m in re.finditer(r'\bDim\s+([A-Za-z0-9_\u3000-\u9fff]+)\b', body):
        info["dim_vars"].append(m.group(1))
    return info


def col_in_sqref(col_letter, sqref):
    """列文字がデータ検証範囲sqrefに含まれるかを判定。"""
    try:
        from openpyxl.utils import column_index_from_string
        col = column_index_from_string(col_letter)
        for rng in sqref.ranges:
            if rng.min_col <= col <= rng.max_col:
                return True
    except Exception:
        pass
    return False


def format_cell(value):
    if value is None:
        return ""
    s = str(value)
    return s.replace("\n", "⏎")


PAPER_SIZES = {0: "既定", 1: "Letter", 8: "A3", 9: "A4", 11: "A4(小)", 5: "Legal", 7: "B4", 12: "B5"}


def header_text(item):
    """HeaderFooter項目(left/center/right)のテキストを集約する。"""
    if item is None:
        return ""
    parts = []
    for pos in ("left", "center", "right"):
        sub = getattr(item, pos, None)
        if sub is not None and sub.text:
            parts.append("%s=%s" % (pos, sub.text))
    return " | ".join(parts)


def print_settings_desc(ws):
    """ページ設定(PageSetup)の要約を返す。帳票仕様の把握に使う。"""
    ps = ws.page_setup
    pm = ws.page_margins
    out = []
    if ps is not None:
        if ps.orientation:
            out.append("向き=%s" % ps.orientation)
        if ps.paperSize:
            out.append("用紙=%s" % PAPER_SIZES.get(ps.paperSize, "コード%s" % ps.paperSize))
        if ps.scale and ps.scale != 100:
            out.append("倍率=%s%%" % ps.scale)
        if ps.fitToWidth:
            out.append("横=%dページ幅に収める" % ps.fitToWidth)
        if ps.fitToHeight:
            out.append("縦=%dページ高さに収める" % ps.fitToHeight)
    if ws.print_area:
        out.append("印刷範囲=%s" % ws.print_area)
    if ws.print_title_rows:
        out.append("タイトル行=%s" % ws.print_title_rows)
    if ws.print_title_cols:
        out.append("タイトル列=%s" % ws.print_title_cols)
    if pm is not None:
        m = []
        for name in ("left", "right", "top", "bottom", "header", "footer"):
            v = getattr(pm, name)
            if v is not None:
                m.append("%s=%s" % (name, v))
        if m:
            out.append("余白: " + ", ".join(m))
    hd = header_text(ws.oddHeader) or header_text(ws.evenHeader)
    ft = header_text(ws.oddFooter) or header_text(ws.evenFooter)
    if hd:
        out.append("ヘッダー: " + hd)
    if ft:
        out.append("フッター: " + ft)
    return out


def dxf_desc(dxf):
    """条件付き書式の書式(DifferentialStyle)を要約する。"""
    if dxf is None:
        return ""
    parts = []
    try:
        f = dxf.font
        if f is not None:
            c = f.color
            if c is not None and getattr(c, "rgb", None):
                parts.append("font=%s" % c.rgb)
            if f.bold:
                parts.append("太字")
    except Exception:
        pass
    try:
        fill = dxf.fill
        if fill is not None:
            c = fill.fgColor
            if c is not None and getattr(c, "rgb", None):
                parts.append("塗り=%s" % c.rgb)
    except Exception:
        pass
    return ",".join(parts)


def conditional_formats(ws):
    """条件付き書式ルールの一覧を返す。色分けが業務ステータスの視覚表現になっていることに注意。"""
    out = []
    try:
        cfs = ws.conditional_formatting
    except Exception:
        return out
    for cf in cfs:
        sqref = str(cf.sqref)
        for rule in cf.rules:
            formula = "; ".join(rule.formula) if rule.formula else ""
            op = getattr(rule, "operator", None) or ""
            try:
                d = dxf_desc(getattr(rule, "dxf", None))
            except Exception:
                d = ""
            out.append({"sqref": sqref, "type": rule.type, "op": op,
                        "formula": formula, "dxf": d})
    return out


def read_app_props(zpath):
    """docProps/app.xml から編集時間・アプリ情報を返す。資産規模把握に使う。"""
    root = read_zip_xml(zpath, "docProps/app.xml")
    if root is None:
        return {}
    out = {}
    for child in root:
        tag = child.tag.split("}")[-1]
        if tag in ("Application", "TotalTime", "Company", "Template"):
            out[tag] = (child.text or "").strip()
    return out


def external_links(zpath):
    """外部リンク（他ブック参照）の対象パス一覧を返す。他システム連携の手がかり。"""
    import zipfile
    import re as _re
    links = []
    try:
        with zipfile.ZipFile(zpath) as z:
            names = set(z.namelist())
            for entry in sorted(names):
                if not _re.match(r"xl/externalLinks/externalLink\d+\.xml$", entry):
                    continue
                try:
                    content = z.read(entry).decode("utf-8", errors="replace")
                except KeyError:
                    continue
                base = _re.match(r"(xl/externalLinks/externalLink\d+)", entry).group(1)
                for t in _re.findall(r'link="([^"]*)"', content):
                    if t not in links:
                        links.append(t)
                rels_name = base + ".rels"
                if rels_name in names:
                    rc = z.read(rels_name).decode("utf-8", errors="replace")
                    for t in _re.findall(r'Target="([^"]*)"', rc):
                        if t not in links:
                            links.append(t)
    except Exception:
        pass
    return links


def editable_columns(ws, sample_rows=8):
    """シート保護時の編集可能（ロック解除）な列を返す。入力可能列の特定に使う。"""
    from openpyxl.utils import get_column_letter
    out = []
    for col in range(1, min(ws.max_column, 30) + 1):
        unlocked = 0
        for r in range(1, min(ws.max_row, sample_rows) + 1):
            cell = ws.cell(row=r, column=col)
            if cell.protection is not None and not cell.protection.locked:
                unlocked += 1
        if unlocked:
            out.append(get_column_letter(col))
    return out


def sheet_hyperlinks(ws):
    """シート上のハイパーリンク一覧（画面遷移・参照の手がかり）。"""
    out = []
    try:
        for ref, hl in ws._hyperlinks.items():
            tgt = hl.target or hl.location
            if tgt:
                out.append((ref, tgt))
    except Exception:
        pass
    return out


def _split_concat(buf):
    """文字列連結式を & で分割する。括弧/クォート内の & は無視（深さ追跡）。"""
    parts, cur = [], []
    depth, i = 0, 0
    while i < len(buf):
        c = buf[i]
        if c == '"':
            cur.append(c)
            i += 1
            while i < len(buf) and buf[i] != '"':
                cur.append(buf[i])
                i += 1
            if i < len(buf):
                cur.append(buf[i])
                i += 1
            continue
        if c == "(":
            depth += 1
        elif c == ")":
            depth = max(0, depth - 1)
        if c == "&" and depth == 0:
            parts.append("".join(cur))
            cur = []
            i += 1
            continue
        cur.append(c)
        i += 1
    if cur:
        parts.append("".join(cur))
    return parts


def _is_literal(seg):
    """セグメント全体が1つの文字列リテラルならTrue。"""
    s = seg.strip()
    if len(s) < 2 or not s.startswith('"'):
        return False
    return s.endswith('"')


def vba_sql_statements(code):
    """sql変数への代入文から文字列連結(& と _改行継続)を再構成してSQLを抽出する。
    変数は `{式}` のプレースホルダとして残す（値は実行時まで不明のため）。"""
    out = []
    lines = code.splitlines()
    n = len(lines)
    i = 0
    while i < n:
        m = re.match(r"^\s*(?:sql|SQL|strSQL|q)\s*=\s*(.+)$", lines[i])
        if not m:
            i += 1
            continue
        buf = m.group(1)
        while buf.rstrip().endswith("& _") and i + 1 < n:
            i += 1
            buf += " " + lines[i]
        joined = ""
        for seg in _split_concat(buf):
            seg = seg.strip()
            if _is_literal(seg):
                joined += seg[1:-1].replace('""', '"')
            elif seg:
                joined += "{%s}" % re.sub(r"\s+", " ", seg)
        joined = joined.strip()
        if joined and re.search(r"\b(SELECT|INSERT|UPDATE|DELETE)\b", joined, re.IGNORECASE):
            out.append(joined)
        i += 1
    return out


def _col_letter(index):
    """0始まりの配列インデックスをシート列文字に変換 (index 0 -> A)。"""
    try:
        from openpyxl.utils import get_column_letter
        return get_column_letter(index + 1)
    except Exception:
        return "?" * (index + 1)


def _strip_vba_comment(s):
    """VBAコメント('以降)を除去する。クォート文字列内の'は対象外。"""
    in_str = False
    for i, ch in enumerate(s):
        if ch == '"':
            in_str = not in_str
        elif ch == "'" and not in_str:
            return s[:i]
    return s


def vba_column_maps(code):
    """4.2技法の自動化: 配列インデックス=シート列(0始まり)から列マッピングを復元する。
    取り込み(in, DB/セル→シート): 二次元配列(w行, N) = ![フィールド] / Range("X"&行) の形
    セット(out, シート→DB):       ![フィールド] = Range("X" & w行) 等の形
    戻り値: (in_rows, out_rows)
      in_rows:  [{"col": "B", "index": 1, "src": "DBフィールド:コード"}, ...]
      out_rows: [{"field": "数量", "src": "シートE列"}, ...]
    """
    in_rows, out_rows = [], []
    for line in code.splitlines():
        s = line.strip()
        m = re.match(r"([A-Za-z0-9_\u3000-\u9fff]+)\((?:w行|行|i|n|カウント),\s*(\d+)\)\s*=\s*(.+)", s)
        if m:
            idx = int(m.group(2))
            rhs = _strip_vba_comment(m.group(3)).strip()
            fld = re.search(r"!\[([^\]]+)\]", rhs)
            if fld:
                src = "DBフィールド:%s" % fld.group(1)
            elif re.search(r"(w行|行|i|n)\s*\+\s*1", rhs):
                src = "連番(行番号+1)"
            elif re.search(r'Range\("([A-Za-z]+)', rhs):
                src = "シート%s列" % re.search(r'Range\("([A-Za-z]+)', rhs).group(1)
            elif re.search(r"\bCells\(", rhs):
                src = "セル値"
            else:
                src = rhs
            in_rows.append({"col": _col_letter(idx), "index": idx, "src": src})
            continue
        m = re.match(r"!\[([^\]]+)\]\s*=\s*(.+)", s)
        if m and not re.match(r"(?:If|ElseIf|While|Select|Case)\b", s):
            field = m.group(1)
            rhs = _strip_vba_comment(m.group(2)).strip()
            if "未入力の場合Null変換" in rhs:
                cm = re.search(r'Range\("([A-Za-z]+)', rhs)
                src = "Null変換(シート%s列)" % cm.group(1) if cm else "Null変換"
            elif re.search(r'Range\("([A-Za-z]+)', rhs):
                src = "シート%s列" % re.search(r'Range\("([A-Za-z]+)', rhs).group(1)
            elif re.search(r"\bNow\b|\bDate\b", rhs):
                src = "現在日付"
            elif re.search(r"\bNull\b", rhs, re.IGNORECASE):
                src = "Null(未設定)"
            elif re.search(r"\bSheets\(", rhs):
                src = "他シート値: %s" % rhs
            else:
                src = rhs
            out_rows.append({"field": field, "src": src})
    return in_rows, out_rows


def commented_out(code):
    """コメントアウトされたコード行（デッドコードの証拠）を返す。
    Sub/Function/イベントハンドラ宣言がコメント中に残っているものを拾う。"""
    out = []
    for m in re.finditer(
        r"^\s*'\s*(?:Public |Private )?(?:Sub|Function|Property)\s+"
        r"([A-Za-z0-9_\u3000-\u9fff]+)",
        code, re.MULTILINE):
        out.append(m.group(1))
    return sorted(set(out))


def analyze_sheet(wb, ws, sheet_buttons, sheet_names, code_names):
    """openpyxlのワークシートから構造詳細をMarkdown文字列で返す。"""
    L = []
    L.append(f"## {ws.title}\n")
    L.append(f"- シート状態: `{ws.sheet_state}`（hidden=非表示, veryHidden=完全非表示）")
    L.append(f"- 使用範囲: `{ws.dimensions}`  (最大行 {ws.max_row} / 最大列 {ws.max_column})")
    L.append(f"- フリーズペイン: `{ws.freeze_panes}`")
    L.append(f"- オートフィルタ: `{ws.auto_filter.ref}`")
    L.append(f"- シート保護: {ws.protection.sheet}")
    if ws.protection.sheet:
        if ws.protection.password:
            L.append("- 保護パスワード: **あり**")
        edit_cols = editable_columns(ws)
        if edit_cols:
            L.append(f"- 編集可能な列（保護中の入力可能列）: {', '.join(edit_cols)}")
        else:
            L.append("- 編集可能なセル: なし（全面保護）")
    L.append(f"- 結合セル: {len(ws.merged_cells.ranges)} 件")
    L.append(f"- タブ色: {ws.sheet_properties.tabColor.rgb if ws.sheet_properties.tabColor else 'なし'}")
    L.append(f"- 印刷範囲: {getattr(ws, 'print_area', None)}")
    L.append("")

    # 列幅
    colw = []
    for col_letter, dim in sorted(ws.column_dimensions.items()):
        if dim.width and dim.width != 0 and isinstance(dim.width, (int, float)):
            colw.append(f"{col_letter}={dim.width:g}")
    if colw:
        L.append(f"**列幅**: {' | '.join(colw)}")
        L.append("")

    # 結合セルと値
    merged_vals = []
    for mr in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
        v = ws.cell(row=mr.min_row, column=mr.min_col).value
        merged_vals.append(f"`{mr.coord}`: {format_cell(v)}")
    if merged_vals:
        L.append("**結合セル**\n")
        for mv in merged_vals:
            L.append(f"- {mv}")
        L.append("")

    # 上位ヘッダー行のダンプ（5行まで）
    dump_rows = min(5, ws.max_row)
    L.append("**上位行の内容**\n")
    for r in range(1, dump_rows + 1):
        cells = []
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and format_cell(v) != "":
                cells.append(f"{ws.cell(row=r, column=c).coordinate}={format_cell(v)}")
        L.append(f"- 行{r}: " + ("; ".join(cells) if cells else "(空)"))
    L.append("")

    # 列ごとの定義（ヘッダー行2行分 + 数式/データ検証/コメント有無）
    L.append("**列一覧**\n")
    L.append("| 列 | ヘッダー行1 | ヘッダー行2 | データ型候補 | 入力規制 | コメント |")
    L.append("|---|---|---|---|---|---|")
    for col_letter in sorted(ws.column_dimensions.keys()):
        ci = ws[col_letter + "1"].column
        h1 = ws.cell(row=1, column=ci).value
        h2 = ws.cell(row=2, column=ci).value
        vals = []
        for r in range(3, min(ws.max_row, 8) + 1):
            v = ws.cell(row=r, column=ci).value
            if v is not None:
                vals.append(str(v))
        dtype = guess_type(h1, h2, vals)
        dv = ws.data_validations
        dvl = []
        for d in dv.dataValidation:
            if col_in_sqref(col_letter, d.sqref):
                dvl.append(d.type or "制限")
        dvl_str = ",".join(x for x in dvl if x) or ""
        comment = None
        for r in range(1, 3):
            cm = ws.cell(row=r, column=ci).comment
            if cm is not None:
                comment = cm.text.replace("\n", " ")
                break
        L.append(f"| {col_letter} | {format_cell(h1)} | {format_cell(h2)} | {dtype} | {dvl_str} | {comment or ''} |")
    L.append("")

    # 数式
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append((cell.coordinate, cell.value))
    if formulas:
        L.append(f"**数式一覧 ({len(formulas)} 件)**\n")
        for coord, f in formulas:
            L.append(f"- `{coord}`: `{f}`")
        L.append("")

    # データ検証の詳細
    dvs = []
    for d in ws.data_validations.dataValidation:
        dvs.append({"type": d.type, "formula1": d.formula1, "formula2": d.formula2,
                    "sqref": d.sqref, "allow_blank": d.allow_blank})
    if dvs:
        L.append("**データ検証（入力規制）**\n")
        for d in dvs:
            L.append(f"- 種類={d['type']} 範囲=`{d['sqref']}` formula1=`{d['formula1']}` formula2=`{d['formula2']}` blankOK={d['allow_blank']}")
        L.append("")

    # 印刷設定（帳票仕様の要）
    print_info = print_settings_desc(ws)
    if print_info:
        L.append("**印刷設定**\n")
        for p in print_info:
            L.append(f"- {p}")
        L.append("")

    # 条件付き書式（色分け = 業務ステータスの視覚表現）
    cfs = conditional_formats(ws)
    if cfs:
        L.append("**条件付き書式**\n")
        L.append("| 範囲 | 種類 | 演算子/条件 | 数式 | 書式 |")
        L.append("|---|---|---|---|---|")
        for c in cfs:
            L.append(f"| {c['sqref']} | {c['type']} | {c['op'] or '-'} | `{c['formula']}` | {c['dxf'] or '-'} |")
        L.append("")

    # ハイパーリンク（画面遷移・他ブック参照の手がかり）
    hls = sheet_hyperlinks(ws)
    if hls:
        L.append("**ハイパーリンク**\n")
        for ref, tgt in hls:
            L.append(f"- `{ref}` → `{tgt}`")
        L.append("")

    # コメント
    comments = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                comments.append((cell.coordinate, cell.comment.text.replace("\n", " ")))
    if comments:
        L.append("**コメント**\n")
        for coord, txt in comments:
            L.append(f"- `{coord}`: {txt}")
        L.append("")

    # ボタン→マクロ
    if sheet_buttons:
        L.append(f"**ボタン → マクロ割当 ({len(sheet_buttons)} 件)**\n")
        for b in sheet_buttons:
            L.append(f"- `{b}`")
        L.append("")

    return "\n".join(L)


def guess_type(h1, h2, vals):
    h = "".join(str(x) for x in (h1, h2) if x)
    if any(k in h for k in ("コード", "NO", "番号", "管理NO", "注番", "棚番")):
        return "文字列(コード)"
    if any(k in h for k in ("日", "納期", "期間")):
        return "日付"
    if any(k in h for k in ("数量", "単価", "金額", "数", "残")):
        return "数値"
    if any(k in h for k in ("者", "氏名", "担当", "取引先", "依頼")):
        return "文字列(名称)"
    if any(k in h for k in ("備考", "型式", "単位")):
        return "文字列"
    if vals and all(re.match(r"^-?\d+(\.\d+)?$", str(v)) for v in vals):
        return "数値"
    return "文字列"


def db_type_name(vals):
    """カラムの値リストから型を推定する。"""
    types = sorted(set(type(v).__name__ for v in vals if v is not None))
    if any(t in ("datetime", "date") for t in types):
        return "日付/時刻"
    if "int" in types and not any(t in ("float", "decimal", "Decimal") for t in types):
        return "整数"
    if any(t in ("float", "decimal", "Decimal") for t in types):
        return "数値"
    if "bool" in types:
        return "真偽"
    if any(t in ("bytes", "bytearray") for t in types):
        return "バイナリ"
    if "str" in types:
        return "文字列"
    return ",".join(types) if types else "-"


def db_sample_vals(vals, limit=3, maxlen=40):
    out = []
    for v in vals:
        if v is None:
            continue
        if isinstance(v, bytes):
            s = "<bytes %d>" % len(v)
        else:
            s = str(v)
        if len(s) > maxlen:
            s = s[:maxlen] + "..."
        out.append(s)
        if len(out) >= limit:
            break
    return out


def db_select_cols(sel):
    """SQLのSELECT列リスト文字列から列名候補を抽出する。"""
    out = []
    for item in sel.split(","):
        item = item.strip()
        if not item:
            continue
        if item == "*":
            out.append("*")
            continue
        item = re.split(r"\s+AS\s+", item, flags=re.IGNORECASE)[0].strip()
        ids = re.findall(r"[A-Za-z0-9_\u3000-\u9fff]+", item)
        if not ids:
            continue
        col = ids[-1]
        if col.upper() in ("DISTINCT", "TOP", "ALL", "PERCENT", "FROM"):
            continue
        out.append(col)
    return out


def collect_vba_db_refs(vba_mods):
    """VBAからDBへの参照(T_/Q_テーブル名, ![カラム], SELECT列)を収集する。"""
    table_refs = Counter()
    col_refs = Counter()
    sql_select = []
    for mname, ftype, code in vba_mods:
        for m in re.finditer(r'["\']([TQ]_[A-Za-z0-9_\u3000-\u9fff]+)', code):
            table_refs[m.group(1)] += 1
        for m in re.finditer(r'!\[([^\]\[]+)\]', code):
            col_refs[m.group(1).strip()] += 1
        for m in re.finditer(r"\bSELECT\s+([^;\"']*?)\s+FROM\s+([A-Za-z0-9_\u3000-\u9fff]+)",
                             code, re.IGNORECASE):
            sql_select.append((mname, m.group(2), db_select_cols(m.group(1))))
    return {"tables": table_refs, "columns": col_refs, "sql": sql_select}


def analyze_access_db(db_path, vba_mods):
    """access_parserでDBスキーマを検証し、VBA参照と突き合わせたMarkdownを返す。"""
    try:
        from access_parser import AccessParser
    except ImportError:
        return None
    db = AccessParser(db_path)
    catalog = db.catalog
    user_tables = [t for t in catalog
                   if not t.startswith("MSys") and not t.startswith("f_")
                   and not t.startswith("MSysObjects")]
    queries = []
    try:
        mo = db.parse_table("MSysObjects")
        for t, n in zip(mo.get("Type", []), mo.get("Name", [])):
            if t == 5:
                queries.append(n)
    except Exception:
        pass
    tables = {}
    for t in sorted(user_tables):
        try:
            tbl = db.parse_table(t)
        except Exception as e:
            tables[t] = {"error": str(e), "columns": {}, "rows": 0}
            continue
        cols = {}
        for c, vals in tbl.items():
            cols[c] = {"type": db_type_name(vals),
                       "nonnull": sum(1 for v in vals if v is not None),
                       "samples": db_sample_vals(vals)}
        tables[t] = {"columns": cols,
                     "rows": max((len(v) for v in tbl.values()), default=0)}
    return {"path": db_path, "tables": tables, "queries": queries,
            "vba": collect_vba_db_refs(vba_mods)}


def write_db_report(out, db_path, vba_mods):
    """DBスキーマ検証レポート 50_db_schema.md を書き出す。"""
    report = analyze_access_db(db_path, vba_mods)
    if report is None:
        with open(os.path.join(out, "50_db_schema.md"), "w", encoding="utf-8") as f:
            f.write("# DBスキーマ検証\n\n")
            f.write("`access_parser` が未導入のためDB解析をスキップしました。\n")
            f.write("`pip install access_parser` で導入後、`--db` を付けて再実行してください。\n")
        return

    tables = report["tables"]
    refs = report["vba"]
    total_rows = sum(t["rows"] for t in tables.values())
    data_rows = sum(tables[t]["rows"] for t in tables if "採番" not in t)
    table_cols = {t: set(d["columns"].keys()) for t, d in tables.items() if "columns" in d}

    L = []
    L.append("# DBスキーマ検証\n")
    L.append("")
    L.append(f"- DBファイル: `{os.path.basename(db_path)}`")
    L.append(f"- ユーザーテーブル数: {len(tables)}")
    L.append(f"- クエリ数: {len(report['queries'])}")
    L.append(f"- 全行数: {total_rows}")
    L.append(f"- 実データ行数(採番テーブル除く): {data_rows}")
    if tables and data_rows == 0:
        L.append("- **判定: 実データ0行の初期テンプレート状態。実運用DBのスキーマと突き合わせが必要**")
    L.append("")

    L.append("## テーブル定義\n")
    if not tables:
        L.append("(ユーザーテーブルなし)\n")
    for t, d in tables.items():
        L.append(f"### {t} ({d['rows']}行)\n")
        if d.get("error"):
            L.append(f"- 解析エラー: {d['error']}\n")
            continue
        L.append("| カラム | 型 | 非NULL数 | サンプル値 |")
        L.append("|---|---|---|---|")
        for c, ci in d["columns"].items():
            L.append(f"| {c} | {ci['type']} | {ci['nonnull']} | {', '.join(ci['samples']) or '-'} |")
        L.append("")

    L.append("## クエリ一覧\n")
    if report["queries"]:
        for q in sorted(report["queries"]):
            L.append(f"- `{q}`")
    else:
        L.append("(クエリなし)")
    L.append("")

    L.append("## VBAとの突き合わせ\n")
    L.append("### VBAが参照するテーブル/クエリ\n")
    referenced = set(refs["tables"].keys())
    for t, cnt in refs["tables"].most_common():
        if t.startswith("Q_"):
            exists = "存在" if t in report["queries"] else "**不在 (C15)**"
        else:
            exists = "存在" if t in tables else "**不在 (C15)**"
        L.append(f"- `{t}` ({cnt}回) → {exists}")
    if not referenced:
        L.append("(VBAからDBテーブル参照なし)")
    L.append("")

    dormant = [t for t in sorted(tables) if t not in referenced]
    L.append("### VBAから未参照のテーブル（休眠機能の手がかり C19）\n")
    if dormant:
        for t in dormant:
            L.append(f"- `{t}`")
    else:
        L.append("(なし)")
    L.append("")

    L.append("### カラム存在チェック（VBAの `![カラム]` 参照）\n")
    L.append("| 参照カラム | 参照回数 | マッチするテーブル |")
    L.append("|---|---|---|")
    for col, cnt in refs["columns"].most_common():
        matched = [t for t, cs in table_cols.items() if col in cs]
        if matched:
            L.append(f"| {col} | {cnt} | {', '.join(sorted(matched))} |")
        else:
            L.append(f"| {col} | {cnt} | **該当なし (C15/C20)** |")
    if not refs["columns"]:
        L.append("| - | - | (参照なし) |")
    L.append("")

    L.append("### SQLのSELECT列と実カラムの比較\n")
    for mname, t, cols in refs["sql"]:
        real = table_cols.get(t)
        if real is None:
            L.append(f"- `{mname}`: `FROM {t}` → **テーブルが実DBに無い (C15)**")
            continue
        missing = [c for c in cols if c != "*" and c not in real]
        if missing:
            L.append(f"- `{mname}`: SELECT ... FROM `{t}` → 実DBに無い列: {missing}")
    if not refs["sql"]:
        L.append("(SELECT文の検出なし)")

    with open(os.path.join(out, "50_db_schema.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")


def make_report(out, wb_path, zpath, db_path=None):
    os.makedirs(out, exist_ok=True)
    sheets_dir = os.path.join(out, "sheets")
    vba_dir = os.path.join(out, "vba")
    os.makedirs(sheets_dir, exist_ok=True)
    os.makedirs(vba_dir, exist_ok=True)

    # ---- ワークブック基本情報 ----
    from openpyxl import load_workbook
    wb = load_workbook(wb_path, data_only=False, keep_vba=True)
    sheets = sheet_order(zpath)
    sheet_entries = ["xl/worksheets/sheet%d.xml" % i for i in range(1, len(sheets) + 1)]
    btn_map = buttons_map(zpath, sheet_entries)
    imgs = collect_images(zpath)

    props = wb.properties
    size = os.path.getsize(wb_path)
    app_props = read_app_props(zpath)
    ext_links = external_links(zpath)

    with open(os.path.join(out, "00_workbook_overview.md"), "w", encoding="utf-8") as f:
        f.write("# ワークブック概要\n\n")
        f.write(f"- ファイル: `{os.path.basename(wb_path)}`\n")
        f.write(f"- サイズ: {size:,} bytes\n")
        f.write(f"- 最終更新: {datetime.fromtimestamp(os.path.getmtime(wb_path)):%Y-%m-%d %H:%M:%S}\n")
        f.write(f"- タイトル: {props.title}\n")
        f.write(f"- 作成者: {props.creator}\n")
        f.write(f"- 最終更新者: {props.lastModifiedBy}\n")
        f.write(f"- 作成日: {props.created}\n")
        f.write(f"- 更新日: {props.modified}\n")
        if "TotalTime" in app_props:
            try:
                hrs = int(app_props["TotalTime"]) / 3600
                f.write(f"- 総編集時間: {hrs:.1f} 時間 (TotalTime)\n")
            except Exception:
                f.write(f"- TotalTime: {app_props['TotalTime']}\n")
        if app_props.get("Application"):
            f.write(f"- 作成アプリ: {app_props['Application']}\n")
        if app_props.get("Company"):
            f.write(f"- 会社: {app_props['Company']}\n")
        if app_props.get("Template"):
            f.write(f"- テンプレート: {app_props['Template']}\n")
        f.write("\n## 外部リンク（他ブック参照）\n\n")
        if ext_links:
            for t in ext_links:
                f.write(f"- `{t}`\n")
        else:
            f.write("(なし)\n")
        f.write("\n## シート構成\n\n")
        f.write("| # | シート名 | sheetId | 状態 | マクロ有無 | ボタン数 |\n")
        f.write("|---|---|---|---|---|---|\n")
        for i, (name, sid, rid) in enumerate(sheets, 1):
            sn = "sheet%d" % i
            btns = btn_map.get(sn, [])
            f.write(f"| {i} | {name} | {sid} | {wb[name].sheet_state} | - | {len(btns)} |\n")
        f.write("\n## 定義名 (Named Range / PrintArea)\n\n")
        names = []
        for n in wb.defined_names:
            d = wb.defined_names[n]
            try:
                for dest in d.destinations:
                    names.append(f"- `{n}` → `{dest[0]}!{dest[1]}`")
            except Exception:
                names.append(f"- `{n}` → `{d.value}`")
        f.write("\n".join(names) + "\n" if names else "(なし)\n")
        f.write("\n## 画像/図形\n\n")
        for k in sorted(imgs):
            f.write(f"- `{k}`: {imgs[k]} 個\n")

    # ---- シート一覧 ----
    with open(os.path.join(out, "10_sheet_list.md"), "w", encoding="utf-8") as f:
        f.write("# シート一覧・役割推定\n\n")
        f.write("| # | シート名 | 状態 | 使用範囲 | フリーズ | オートフィルタ | 保護 | ボタン | 数式 | 役割推定 |\n")
        f.write("|---|---|---|---|---|---|---|---|---|---|\n")
        for i, (name, sid, rid) in enumerate(sheets, 1):
            ws = wb[name]
            sn = "sheet%d" % i
            btns = btn_map.get(sn, [])
            fcount = sum(1 for row in ws.iter_rows()
                         for cell in row if isinstance(cell.value, str) and cell.value.startswith("="))
            role = guess_role(ws, btns, name)
            state = ws.sheet_state
            if state != "visible":
                role = "非表示(%s)" % state
            f.write(f"| {i} | {name} | {state} | {ws.dimensions} | {ws.freeze_panes} | {ws.auto_filter.ref} | "
                    f"{ws.protection.sheet} | {len(btns)} | {fcount} | {role} |\n")

        hidden = []
        for n, _, _ in sheets:
            if wb[n].sheet_state != "visible":
                hidden.append((n, wb[n].sheet_state))
        if hidden:
            f.write("\n## 非表示シート\n\n")
            for n, st in hidden:
                f.write(f"- `{n}` ({st})\n")
            f.write("\n非表示シートは参照用・過去バージョンの残骸の可能性がある。VBAからの参照有無を必ず確認する。\n")

    # ---- VBA ----
    vba_mods = extract_vba(wb_path)
    proc_names = set()
    module_procs = {}
    for mname, ftype, code in vba_mods:
        procs = parse_procs(code)
        for p in procs:
            name = p["signature"][2]
            proc_names.add(name)
        module_procs[mname] = (ftype, procs, code)

    code_names = [m for m, _, _ in vba_mods if m not in ("ThisWorkbook",)]
    sheet_names = [s for s, _, _ in sheets]

    # モジュールごとの完全ソース保存
    for mname, ftype, code in vba_mods:
        safe = re.sub(r"[^\w\-]", "_", mname)
        with open(os.path.join(vba_dir, "%s.%s.txt" % (safe, ftype or "txt")), "w", encoding="utf-8") as f:
            f.write(code)

    # ---- VBAサマリ ----
    with open(os.path.join(out, "20_vba_summary.md"), "w", encoding="utf-8") as f:
        f.write("# VBA解析サマリ\n\n")
        f.write(f"## モジュール一覧 ({len(vba_mods)} 件)\n\n")
        f.write("| モジュール | タイプ | 行数 | プロシージャ数 |\n|---|---|---|---|\n")
        for mname, ftype, code in vba_mods:
            f.write(f"| {mname} | {ftype or 'module'} | {len(code.splitlines())} | {len(module_procs[mname][1])} |\n")

        f.write("\n## プロシージャ一覧\n\n")
        f.write("| モジュール | プロシージャ | 種別 | 引数 | 行数 | 説明 |\n|---|---|---|---|---|---|\n")
        for mname, ftype, code in vba_mods:
            for p in module_procs[mname][1]:
                vis, kind, name, args = p["signature"]
                doc = first_doc_comment(p["body"], "")
                f.write(f"| {mname} | {name} | {kind} | {args or '-'} | {p['line_count']} | {doc} |\n")

        f.write("\n## 呼び出し関係\n\n")
        f.write("(形式: 呼び出し側モジュール → 呼び出し先プロシージャ)\n\n")
        for mname, ftype, code in vba_mods:
            for p in module_procs[mname][1]:
                info = analyze_proc(p, proc_names, sheet_names, code_names)
                if info["calls"]:
                    f.write(f"- `{mname}.{p['signature'][2]}` → " +
                            ", ".join(sorted(info["calls"])) + "\n")

        f.write("\n## 未解決の呼び出し（外部・Excel関数等）\n\n")
        unresolved = Counter()
        for mname, ftype, code in vba_mods:
            for p in module_procs[mname][1]:
                info = analyze_proc(p, proc_names, sheet_names, code_names)
                for c in info["calls"]:
                    if c not in proc_names:
                        unresolved[c] += 1
        for c, cnt in unresolved.most_common():
            f.write(f"- `{c}` ({cnt}回)\n")

        # グローバル変数・定数
        f.write("\n## グローバル宣言 (Public/モジュールレベル)\n\n")
        for mname, ftype, code in vba_mods:
            for ln in code.splitlines():
                s = ln.strip()
                if re.match(r"^(Public|Private)\s+(Const\s+|Dim\s+|[A-Za-z_])", s) and "As" in s:
                    f.write(f"- `{mname}`: `{s}`\n")
        f.write("\n## 定数\n\n")
        for mname, ftype, code in vba_mods:
            for m in re.finditer(r"(?:Public|Private)?\s*Const\s+([A-Za-z0-9_\u3000-\u9fff]+)\s*=\s*(.+)", code):
                f.write(f"- `{mname}`: `{m.group(1)} = {m.group(2).strip()}`\n")

        f.write("\n## DBテーブル参照\n\n")
        tables = Counter()
        for mname, ftype, code in vba_mods:
            for m in re.finditer(r'["\']([TQ]_[A-Za-z0-9_\u3000-\u9fff]+)', code):
                tables[m.group(1)] += 1
        for t, cnt in tables.most_common():
            f.write(f"- `{t}` ({cnt}回)\n")

        # エラー処理方針 (On Error)
        f.write("\n## エラー処理方針 (On Error)\n\n")
        err_counts = Counter()
        err_snippets = []
        for mname, ftype, code in vba_mods:
            for m in re.finditer(r"On Error\s+(GoTo\s+\w+|Resume Next|GoTo 0)", code, re.IGNORECASE):
                err_counts[m.group(1)] += 1
            for m in re.finditer(r"On Error\s+(GoTo\s+\w+)", code, re.IGNORECASE):
                lbl = m.group(1).split()[1]
                end = code.find("Exit Sub", m.end())
                end = end if end != -1 else code.find("Exit Function", m.end())
                end = end if end != -1 else len(code)
                seg = code[m.end():end]
                if re.search(r"MsgBox", seg, re.IGNORECASE):
                    err_snippets.append((mname, lbl, "MsgBox表示"))
                elif re.search(r"Resume", seg, re.IGNORECASE):
                    err_snippets.append((mname, lbl, "Resume継続"))
                else:
                    err_snippets.append((mname, lbl, "無視/その他"))
        for e, c in err_counts.most_common():
            f.write(f"- `On Error {e}` ... {c}箇所\n")
        if err_snippets:
            f.write("\nハンドラ内の処理傾向:\n")
            for mname, lbl, how in sorted(set(err_snippets)):
                f.write(f"- `{mname}` ({lbl}) → {how}\n")
        if not err_counts:
            f.write("(On Error なし)\n")

        # コメントアウトされたプロシージャ（デッドコードの証拠）
        f.write("\n## コメントアウトされたプロシージャ（デッドコード候補）\n\n")
        commented = False
        for mname, ftype, code in vba_mods:
            names = commented_out(code)
            if names:
                commented = True
                f.write(f"- `{mname}`: {', '.join(names)}\n")
        if not commented:
            f.write("(なし)\n")

        f.write("\n## メッセージ一覧 (MsgBox)\n\n")
        msgs = []
        for mname, ftype, code in vba_mods:
            for m in re.finditer(r'MsgBox\s+"([^"]+)"', code):
                msgs.append(m.group(1))
        for m in sorted(set(msgs)):
            f.write(f"- \"{m}\"\n")

        f.write("\n## UserForm イベントハンドラ（コントロール推測）\n\n")
        for mname, ftype, code in vba_mods:
            if ftype != "frm":
                continue
            controls = set()
            for p in module_procs[mname][1]:
                name = p["signature"][2]
                for suf in EVENT_SUFFIXES:
                    if name.endswith("_" + suf) and not name.startswith("UserForm_"):
                        controls.add(name.split("_" + suf)[0])
            f.write(f"### {mname}\n")
            if controls:
                for c in sorted(controls):
                    f.write(f"- コントロール: `{c}`\n")
            else:
                f.write("- (イベントハンドラなし)\n")

    # ---- フォーム解析 ----
    streams = vba_project_streams(zpath)
    forms = []
    for mname, ftype, code in vba_mods:
        if ftype != "frm":
            continue
        dstr = []
        base = mname.rsplit(".", 1)[0]
        if base + "/f" in streams:
            dstr = form_designer_strings(streams[base + "/f"])
        forms.append((mname, code, dstr))
    with open(os.path.join(out, "25_forms.md"), "w", encoding="utf-8") as f:
        f.write("# UserForm フォーム仕様\n\n")
        if not forms:
            f.write("(UserFormなし)\n")
        for mname, code, dstr in forms:
            f.write(f"## {mname}\n\n")
            info = analyze_form(mname, code, dstr, proc_names, sheet_names, code_names)
            labels = sorted(s for s in dstr if re.fullmatch(r"Label\d+", s))
            extra_jp = sorted(s for s in dstr if s in info["controls"])
            if labels:
                f.write(f"**デザイナーの自動生成ラベル**: {', '.join(labels)}\n\n")
            if extra_jp:
                f.write(f"**デザイナーで検出された名称**: {', '.join(extra_jp)}\n\n")
            f.write("| コントロール | 推測種別 | イベント | コード上で利用するプロパティ |\n")
            f.write("|---|---|---|---|\n")
            for ctrl in sorted(info["controls"]):
                c = info["controls"][ctrl]
                f.write(f"| {ctrl} | {c['type']} | {','.join(c['events']) or '-'} | {','.join(c['usage']) or '-'} |\n")
            if info["form_procs"]:
                f.write("\n**フォーム固有プロシージャ**: " + ", ".join(info["form_procs"]) + "\n")
            f.write("\n")

    # ---- ボタン一覧 ----
    with open(os.path.join(out, "30_buttons.md"), "w", encoding="utf-8") as f:
        f.write("# ボタン → マクロ割当一覧\n\n")
        f.write("| # | シート | マクロ名 | 定義モジュール |\n|---|---|---|---|\n")
        for i, (name, sid, rid) in enumerate(sheets, 1):
            sn = "sheet%d" % i
            for b in btn_map.get(sn, []):
                mod = ""
                for mname, ftype, code in vba_mods:
                    if re.search(r"^\s*Sub\s+" + re.escape(b) + r"\b", code, re.MULTILINE):
                        mod = mname
                        break
                f.write(f"| {i} | {name} | `{b}` | {mod} |\n")

    # ---- クロス参照 ----
    with open(os.path.join(out, "40_cross_references.md"), "w", encoding="utf-8") as f:
        f.write("# クロス参照（VBA ↔ シート / DB）\n\n")
        f.write("## VBAから参照されているシート\n\n")
        refs = defaultdict(set)
        for mname, ftype, code in vba_mods:
            for p in module_procs[mname][1]:
                info = analyze_proc(p, proc_names, sheet_names, code_names)
                for sh in info["sheets"]:
                    refs[sh].add(mname + "." + p["signature"][2])
        for sh in sorted(refs):
            f.write(f"### {sh}\n")
            for r in sorted(refs[sh]):
                f.write(f"- `{r}`\n")

        f.write("\n## VBAから参照されているセル\n\n")
        for mname, ftype, code in vba_mods:
            for p in module_procs[mname][1]:
                info = analyze_proc(p, proc_names, sheet_names, code_names)
                if info["ranges"]:
                    f.write(f"- `{mname}.{p['signature'][2]}`: " +
                            ", ".join(sorted(set(info["ranges"]))) + "\n")

        f.write("\n## 列マッピング候補（配列インデックス = シート列 0始まり）\n\n")
        f.write("`二次元配列(w行, N) = ![フィールド]` のNとシート列(A=0,B=1,...)の対応から、")
        f.write("「DBカラム ↔ シート列」のマップを復元する（取り込み=DB/セル→シート、セット=シート→DB）。\n")
        f.write("同一列の取り込み・セットの対応を突き合わせ、シート列の意味を確定する。\n\n")
        for mname, ftype, code in vba_mods:
            for p in module_procs[mname][1]:
                in_rows, out_rows = vba_column_maps("\n".join(p["body"]))
                if not in_rows and not out_rows:
                    continue
                f.write(f"### {mname}.{p['signature'][2]}\n")
                if in_rows:
                    f.write("\n| シート列 | 配列index | 取得元(取り込み: DB/セル→シート) |\n|---|---|---|\n")
                    for r in sorted(in_rows, key=lambda x: x["index"]):
                        f.write(f"| {r['col']} | {r['index']} | {r['src']} |\n")
                if out_rows:
                    f.write("\n| DBフィールド | 書き込み元(セット: シート→DB) |\n|---|---|\n")
                    for r in out_rows:
                        f.write(f"| {r['field']} | {r['src']} |\n")
                f.write("\n")

        f.write("\n## シート間の数式参照\n\n")
        for i, (name, sid, rid) in enumerate(sheets, 1):
            ws = wb[name]
            hit = False
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        for other, _, _ in sheets:
                            if other != name and other in cell.value:
                                if not hit:
                                    f.write(f"### {name}\n")
                                    hit = True
                                f.write(f"- `{cell.coordinate}`: `{cell.value}` → {other}\n")
            if not hit:
                pass
        f.write("(数式によるシート間参照なし)\n")

        # DBアクセスロジック
        f.write("\n## DBアクセスパターン\n\n")
        for mname, ftype, code in vba_mods:
            for m in re.finditer(r'OpenRecordset\("([^"]+)"', code):
                f.write(f"- `{mname}`: OpenRecordset(\"{m.group(1)}\")\n")
            for m in re.finditer(r'OpenDatabase\(([^\)]+)\)', code):
                f.write(f"- `{mname}`: OpenDatabase({m.group(1)})\n")
            for m in re.finditer(r'cn\.Execute\s*\(?\s*"([^"]+)"', code):
                f.write(f"- `{mname}`: cn.Execute(\"{m.group(1)}\")\n")

        # SQL文の再構成（文字列連結 & と改行継続 _ を復元）
        f.write("\n## SQL文の再構成\n\n")
        f.write("`sql = \"...\" & var & ...` 形式の文字列連結SQLを復元したもの。\n")
        f.write("業務テーブルの取得・更新・抽出条件の理解に使う。\n\n")
        for mname, ftype, code in vba_mods:
            stmts = vba_sql_statements(code)
            for s in stmts:
                first = s.splitlines()[0] if s else ""
                kw = "UPDATE" if re.search(r"\bUPDATE\b", s, re.I) else (
                    "INSERT" if re.search(r"\bINSERT\b", s, re.I) else (
                    "DELETE" if re.search(r"\bDELETE\b", s, re.I) else (
                    "SELECT" if re.search(r"\bSELECT\b", s, re.I) else "その他")))
                f.write(f"- `{mname}` [{kw}] `{first}`\n")

    # ---- シート詳細 ----
    for i, (name, sid, rid) in enumerate(sheets, 1):
        sn = "sheet%d" % i
        ws = wb[name]
        content = analyze_sheet(wb, ws, btn_map.get(sn, []), sheet_names, code_names)
        safe = re.sub(r"[^\w\-]", "_", name)
        with open(os.path.join(sheets_dir, "%02d_%s.md" % (i, safe)), "w", encoding="utf-8") as f:
            f.write(content)

    # ---- DBスキーマ検証 ----
    if db_path:
        log("DBスキーマを解析中: %s" % db_path)
        write_db_report(out, db_path, vba_mods)

    wb.close()
    return True


def guess_role(ws, btns, name):
    if btns and ws.max_row <= 25 and ws.max_column <= 20:
        return "メニュー/操作画面"
    if ws.auto_filter.ref and ws.freeze_panes:
        if "マスタ" in name:
            return "マスタデータ一覧"
        if any(k in name for k in ("確認", "状況")):
            return "照会/確認一覧"
        return "データ入力/一覧表"
    if "審査" in name or "承認" in name:
        return "帳票/承認欄"
    return "その他"


def main():
    ap = argparse.ArgumentParser(description="xlsm解析・中間成果物抽出")
    ap.add_argument("xlsm", help="解析対象の.xlsmファイル")
    ap.add_argument("-o", "--out", default=".", help="出力先ディレクトリ")
    ap.add_argument("--db", metavar="DB", default=None,
                    help="Access DB (.accdb/.mdb) のパス。スキーマ検証レポート(50_db_schema.md)を出力")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsm):
        log("エラー: ファイルがありません: %s" % args.xlsm)
        sys.exit(1)
    if not args.xlsm.lower().endswith((".xlsm", ".xlam", ".xlsb")):
        log("注意: 対象が.xlsm以外です。動作は保証されません。")
    if args.db and not os.path.isfile(args.db):
        log("エラー: DBファイルがありません: %s" % args.db)
        sys.exit(1)
    try:
        import openpyxl  # noqa
    except ImportError:
        log("エラー: openpyxlが必要です。 `pip install openpyxl`")
        sys.exit(1)

    warnings.filterwarnings("ignore")
    ok = make_report(args.out, args.xlsm, args.xlsm, args.db)
    if ok:
        log("完了。出力: %s" % os.path.abspath(args.out))
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
